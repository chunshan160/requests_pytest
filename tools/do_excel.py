#!/usr/bin/env python
# -*- coding:utf-8 -*-
# @Time :2020/4/9 9:55
# @Author :春衫
# @File :do_excel.py

from tools.project_path import *
from openpyxl import load_workbook
from tools.read_config import ReadConfig


class DoExcel:

    @classmethod
    def getCaseDataFromExcel(cls, file_name, sheet_name):

        wb = load_workbook(file_name)
        mode = ReadConfig().read_config(case_config_path)  # 配置文件的内容 字典
        test_data = []
        # for key in mode:  # 遍历这个存在配置文件里面的字典 也就是表单名
        sheet = wb[sheet_name]  # 打开Excel里的这个表单
        if mode[sheet_name] == ['all']:  # 判断value
            for i in range(2, sheet.max_row + 1):
                row_data = cls.row_data(sheet, i, sheet_name)
                test_data.append(row_data)
        elif mode[sheet_name] == []:
            pass
        else:
            for case_id in mode[sheet_name]:
                row_data = cls.row_data(sheet, case_id + 1, sheet_name)
                test_data.append(row_data)
        return test_data

    @classmethod
    def row_data(cls, sheet, i, key):
        row_data = {}
        # 序号
        row_data['case_id'] = sheet.cell(i, 1).value
        # 接口模块
        row_data['interface'] = sheet.cell(i, 2).value
        # 用例标题
        row_data['title'] = sheet.cell(i, 3).value
        # 请求头
        row_data['request_header'] = sheet.cell(i, 4).value
        # 请求方式
        row_data['method'] = sheet.cell(i, 5).value
        # 接口地址
        row_data['url'] = sheet.cell(i, 6).value
        # 参数输入
        row_data['input_params'] = sheet.cell(i, 7).value
        # 期望返回结果
        row_data['expected'] = sheet.cell(i, 8).value
        # 数据库校验
        row_data['check_SQL'] = sheet.cell(i, 9).value
        # 表单名
        row_data['sheet_name'] = key
        return row_data

    @staticmethod
    def write_back(file_name, sheet_name, i, result, TestResult):  # 专门写回数据  i行号  result结果
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]  # 选择表单
        sheet.cell(i, 7).value = result
        sheet.cell(i, 8).value = TestResult
        wb.save(file_name)  # 保存结果

    @classmethod
    def updata_tel(cls, tel, file_name, sheet_name):
        '''更新Excel表单里的手机号'''
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(2, 1).value = tel
        wb.save(file_name)


if __name__ == '__main__':
    test_data = DoExcel.getCaseDataFromExcel(test_data_path, "register")
    print(test_data)
